from common.db_ops import generate_m00_dict
from common.lib import encode_to_utf8, get_project_root
from common.memory import MemWriter
from common.translate import convert_into_eng
from json import dumps
from openpyxl import load_workbook

import os
import sqlite3
import sys


class GetPlayer:

    writer = None

    def __init__(self, address, debug=False):
        if not GetPlayer.writer:
            GetPlayer.writer = MemWriter()

        if debug:
            self.address = address
        else:
            self.address = GetPlayer.writer.unpack_to_int(address)

        self.ja_player_name = GetPlayer.writer.read_string(self.address + 24)
        self.en_player_name = self.__get_en_player_name(name=self.ja_player_name)
        self.ja_sibling_name = GetPlayer.writer.read_string(self.address + 96)
        self.en_sibling_name = self.__get_en_player_name(name=self.ja_sibling_name)
        self.sibling_relationship = self.__determine_sibling_relationship()

        self.__write_player()
        self.__load_dialog_into_db()


    def __determine_sibling_relationship(self):
        check_byte = GetPlayer.writer.read_bytes(self.address + 96 + 19, size=1)
        if check_byte == b"\x01":
            return "older_brother"
        if check_byte == b"\x02":
            return "younger_brother"
        if check_byte == b"\x03":
            return "older_sister"
        if check_byte == b"\x04":
            return "younger_sister"


    def __get_en_player_name(self, name: str):
        player_names = generate_m00_dict(files="'custom_player_names'")

        if name in player_names:
            return player_names[name]

        return convert_into_eng(word=name)


    def __write_player(self):
        db_file = get_project_root("misc_files/clarity_dialog.db")

        query = f"""
        BEGIN TRANSACTION;
        DELETE FROM player;
        INSERT INTO player (type, name) VALUES
            ('player', '{self.ja_player_name}'),
            ('sibling', '{self.ja_sibling_name}'),
            ('sibling_relationship', '{self.sibling_relationship}');
        END TRANSACTION;
        """

        try:
            conn = sqlite3.connect(db_file)
            cursor = conn.cursor()
            cursor.executescript(query)
            conn.commit()
        finally:
            conn.close()


    def __replace_with_en_names(self, string: str):
        new_string = string.replace("<pnplacehold>", self.en_player_name)
        new_string = new_string.replace("<snplacehold>", self.en_sibling_name)

        if self.sibling_relationship in ["older_brother", "younger_brother"]:
            new_string = new_string.replace("<kyodai_rel1>", "brother")
            new_string = new_string.replace("<kyodai_rel2>", "brother")
            new_string = new_string.replace("<kyodai_rel3>", "brother")
        elif self.sibling_relationship in ["older_sister", "younger_sister"]:
            new_string = new_string.replace("<kyodai_rel1>", "sister")
            new_string = new_string.replace("<kyodai_rel2>", "sister")
            new_string = new_string.replace("<kyodai_rel3>", "sister")

        return new_string


    def __replace_with_ja_names(self, string: str):
        new_string = string.replace("<pnplacehold>", self.ja_player_name)
        new_string = new_string.replace("<snplacehold>", self.ja_sibling_name)

        if self.sibling_relationship == "older_brother":
            new_string = new_string.replace("<kyodai_rel1>", "兄ちゃん")
            new_string = new_string.replace("<kyodai_rel2>", "お兄ちゃん")
            new_string = new_string.replace("<kyodai_rel3>", "兄")
        elif self.sibling_relationship == "younger_brother":
            new_string = new_string.replace("<kyodai_rel1>", "弟")
            new_string = new_string.replace("<kyodai_rel2>", "弟")
            new_string = new_string.replace("<kyodai_rel3>", "弟")
        elif self.sibling_relationship == "older_sister":
            new_string = new_string.replace("<kyodai_rel1>", "姉ちゃん")
            new_string = new_string.replace("<kyodai_rel2>", "お姉ちゃん")
            new_string = new_string.replace("<kyodai_rel3>", "姉")
        elif self.sibling_relationship == "younger_sister":
            new_string = new_string.replace("<kyodai_rel1>", "妹")
            new_string = new_string.replace("<kyodai_rel2>", "妹")
            new_string = new_string.replace("<kyodai_rel3>", "妹")

        return new_string


    def __load_dialog_into_db(self):
        merge_file = get_project_root("misc_files/merge.xlsx")
        db_file = get_project_root("misc_files/clarity_dialog.db")

        #if os.path.exists(merge_file) and os.path.exists(db_file):
        workbook = load_workbook(merge_file)
        worksheet = workbook["Story So Far"]

        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()

        query = "DELETE FROM story_so_far"
        cursor.execute(query)

        for row_num in range(2, worksheet.max_row + 1):
            ja_text = self.__replace_with_ja_names(
                worksheet.cell(row=row_num, column=1).value.replace("'", "''"))

            # if data in fixed english translation column, use it
            if fixed_text := worksheet.cell(row=row_num, column=3).value:
                en_text = self.__replace_with_en_names(fixed_text.replace("'", "''"))
            elif deepl_text := worksheet.cell(row=row_num, column=2).value:
                en_text = self.__replace_with_en_names(deepl_text.replace("'", "''"))
            else:
                # no entry for either type. just use the japanese
                en_text = ja_text

            query = f"INSERT INTO story_so_far (ja, en) VALUES ('{ja_text}', '{en_text}')"
            cursor.execute(query)

        conn.commit()
        conn.close()


def player_name_shellcode(eax_address: int) -> str:

    local_paths = dumps(sys.path).replace("\\", "\\\\")
    log_path = os.path.join(os.path.abspath('.'), 'logs\\console.log').replace("\\", "\\\\")

    shellcode = rf"""
try:
    import sys
    import traceback
    sys.path = {local_paths}
    from hooking.player import GetPlayer
    GetPlayer({eax_address})
except Exception as e:
    with open("{log_path}", "a+") as f:
        f.write(str(traceback.format_exc()))
    """

    return encode_to_utf8(shellcode).decode()
