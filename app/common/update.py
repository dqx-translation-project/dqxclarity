import ast
import json
import os
import re
import requests
import sys
import time
import urllib3
import winreg
from common.constants import (
    GITHUB_CLARITY_CUTSCENE_JSON_URL,
    GITHUB_CLARITY_ITEMS_JSON_URL,
    GITHUB_CLARITY_KEY_ITEMS_JSON_URL,
    GITHUB_CLARITY_MONSTERS_JSON_URL,
    GITHUB_CLARITY_NPC_JSON_URL,
    GITHUB_CLARITY_QUESTS_REQUESTS_JSON_URL,
    GITHUB_CLARITY_VERSION_UPDATE_URL,
    GITHUB_CUSTOM_TRANSLATIONS_ZIP_URL,
)
from common.db_ops import db_query
from common.lib import get_project_root
from io import BytesIO
from loguru import logger as log
from openpyxl import load_workbook
from subprocess import Popen
from zipfile import ZipFile


def download_custom_files() -> None:
    log.info("Downloading custom translation files from dqx-translation-project/dqx-custom-translations.")
    response = download_file(GITHUB_CUSTOM_TRANSLATIONS_ZIP_URL)

    ignore_data = ""

    db_query("DELETE FROM m00_strings")
    zfile = ZipFile(BytesIO(response.content))
    for obj in zfile.infolist():
        if obj.filename.endswith("/"):  # directory
            continue

        if "/json/" in obj.filename and obj.filename.endswith(".json"):
            with zfile.open(obj.filename, "r") as f:
                data = f.read()

            # modify file path to just the name of the file without the extension
            filename = obj.filename.split("/")[-1].rsplit(".", 1)[0]
            read_custom_json_and_import(name=filename, data=data)

        if "/csv/" in obj.filename:
            if obj.filename.endswith("merge.xlsx"):
                with zfile.open(obj.filename, "r") as f:
                    data = f.read()

                read_xlsx_and_import(data)

            if obj.filename.endswith("glossary.csv"):
                with zfile.open(obj.filename, "r") as f:
                    data = f.read()

                read_glossary_and_import(data)

        # ignore.py exists in the remote repo. we'll purge the strings found in this file
        # from the m00_strings table.
        if "/generate_glossary/" in obj.filename and obj.filename.endswith("ignore.py"):
            with zfile.open(obj.filename, "r") as f:
                ignore_data = f.read()

    log.info("Downloading translation files from dqx-translation-project/dqx_translations.")

    # dqx_translations is roughly 17MB~ right now. we only need these files from that repository.
    url_to_db = {
        GITHUB_CLARITY_MONSTERS_JSON_URL: "monsters",
        GITHUB_CLARITY_NPC_JSON_URL: "npcs",
        GITHUB_CLARITY_ITEMS_JSON_URL: "items",
        GITHUB_CLARITY_KEY_ITEMS_JSON_URL: "key_items",
        GITHUB_CLARITY_QUESTS_REQUESTS_JSON_URL: "quests",
        GITHUB_CLARITY_CUTSCENE_JSON_URL: "story_names",
    }

    for url in url_to_db:
        response = download_file(url)
        read_custom_json_and_import(name=url_to_db[url], data=response.content)

    # we want to remove any strings we imported from ignore.py in this repo.
    # these were not included in the glossary from dqx-custom-translations and we do not
    # want them in our m00_strings table either. ordering matters as what's in the url_to_db
    # dict has strings we want to remove afterwards.
    if ignore_data:
        decoded = ignore_data.decode("utf-8")
        match = re.search(r"IGNORE\s*=\s*(\{.*?\})", decoded, re.DOTALL)
        if not match:
            log.warning("Did not parse ignore.py. Ignoring contents.")
            return

    dict_str = match.group(1)
    ignore_dict = ast.literal_eval(dict_str)
    keys_to_remove = list(ignore_dict.keys())

    db_query("DELETE FROM m00_strings WHERE ja IN ({})".format(",".join(f"'{key}'" for key in keys_to_remove)))


def read_custom_json_and_import(name: str, data: str) -> None:
    content = json.loads(data)
    query_list = []

    for item in content:
        key, value = list(content[item].items())[0]

        escaped_value = value.replace("'", "''")
        query_value = f"('{key}', '{escaped_value}', '{name}')"
        query_list.append(query_value)

    insert_values = ",".join(query_list)
    query = f"INSERT INTO m00_strings (ja, en, file) VALUES {insert_values};"
    db_query(query)


def _get_system_python() -> str:
    """Return the system Python executable path from the registry.

    updater.py deletes the venv as part of the update process. launching it
    with sys.executable (the venv Python) causes Windows to lock the
    executable, making shutil.rmtree silently fail to remove it. using the
    system Python avoids this — updater.py only uses stdlib so it runs fine
    without the venv.
    """
    key = r"SOFTWARE\WOW6432Node\Python\PythonCore\3.11-32\InstallPath"
    with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, key) as k:
        return winreg.QueryValueEx(k, "ExecutablePath")[0]


def check_for_updates(update: bool, test_release: str = None) -> None:
    """Checks to see if Clarity is running the latest version of itself. If
    not, will launch updater.py and exit.

    :param update: Whether or not to update after checking for updates.
    :param test_release: If set, skip version comparison and force-install this
        specific release tag (e.g. 'v1.2.3'). For pre-release testing only.
    """
    if not os.path.exists("version.update"):
        log.warning("Couldn't determine current version of dqxclarity. Running as is.")
        return

    with open("version.update") as file:
        cur_ver = file.read().strip()

    if test_release is not None:
        tag_arg = test_release if test_release.startswith("v") else f"v{test_release}"
        log.warning(f"Test mode: force-installing release {tag_arg}.")
        release_url = f"https://api.github.com/repos/dqx-translation-project/dqxclarity/releases/tags/{tag_arg}"
    else:
        log.info("Checking dqxclarity repo for updates...")
        release_url = GITHUB_CLARITY_VERSION_UPDATE_URL

    github_request = download_file(release_url)

    try:
        release = github_request.json()
        tag = release["tag_name"]
        new_ver = tag[1:]

        if test_release is None:
            if new_ver == cur_ver:
                log.success(f"Up to date. Version: {cur_ver}")
                return
            log.warning(f"Out of date! {cur_ver} -> {new_ver}")

        if not update:
            return

        # Download the latest updater.py from the new release tag before running it,
        # so changes to the updater itself are always picked up regardless of what
        # version the user is upgrading from.
        log.info("Grabbing latest updater.")
        updater_url = f"https://raw.githubusercontent.com/dqx-translation-project/dqxclarity/refs/tags/{tag}/app/updater.py"
        response = download_file(updater_url)
        with open("updater.py", "w+b") as f:
            f.write(response.content)

        cmd = [_get_system_python(), "./updater.py"]
        if test_release is not None:
            cmd += ["--release", tag]

        log.info("Launching updater.")
        Popen(cmd)
        sys.exit()

    except Exception as e:
        log.warning(f"There was a problem trying to update. Clarity will continue without updating.\n{e}")
        return


def read_xlsx_and_import(data: str) -> None:
    """We manage a file outside of this repository called merge.xlsx in dqx-
    custom-translations.

    This interacts with that file by reading entries that have been
    manually translated to fix bugs that machine translation introduced
    and inserts/updates the user's local database with override entries
    from the xlsx file.
    """
    workbook = load_workbook(BytesIO(data))
    ws_dialogue = workbook["Dialogue"]
    ws_walkthrough = workbook["Walkthrough"]
    ws_quests = workbook["Quests"]
    ws_story_so_far = workbook["Story So Far"]

    # Dialogue worksheet
    db_query("DELETE FROM fixed_dialog_template")

    values = []
    for i, _ in enumerate(ws_dialogue, start=2):
        source_text = ws_dialogue.cell(row=i, column=1).value
        en_text = ws_dialogue.cell(row=i, column=3).value
        notes = ws_dialogue.cell(row=i, column=4).value
        original_bad_string_text = ws_dialogue.cell(row=i, column=5).value

        if source_text and en_text:
            # bad_string means the machine translation that was returned ended up breaking the game or providing
            # a confusing experience. initially, what was logged in the spreadsheet was the partial string with
            # the player info taken out. we can't use that when we see the text, so we have to iterate over every
            # bad string match we have to find a match. because of how the text was collected before, this exists.
            # when we get rid of all of the partial strings and get the original source strings, we can remove this
            # logic.
            bad_string = 0
            if notes and "BAD STRING" in notes:
                if not original_bad_string_text:
                    bad_string = 1
                else:
                    source_text = original_bad_string_text
            escaped_text = en_text.replace("'", "''")
            values.append(f"('{source_text}', '{escaped_text}', {bad_string})")

    insert_values = ",".join(values)
    query = f"INSERT OR REPLACE INTO fixed_dialog_template (ja, en, bad_string) VALUES {insert_values};"

    db_query(query)

    # Walkthrough worksheet
    values = []
    for i, _ in enumerate(ws_walkthrough, start=2):
        source_text = ws_walkthrough.cell(row=i, column=1).value
        en_text = ws_walkthrough.cell(row=i, column=3).value

        if source_text and en_text:
            escaped_text = en_text.replace("'", "''")
            values.append(f"('{source_text}', '{escaped_text}')")

    insert_values = ",".join(values)
    query = f"INSERT OR REPLACE INTO walkthrough (ja, en) VALUES {insert_values};"

    db_query(query)

    # Quests worksheet
    values = []
    for i, _ in enumerate(ws_quests, start=2):
        source_text = ws_quests.cell(row=i, column=1).value
        en_text = ws_quests.cell(row=i, column=3).value

        if source_text and en_text:
            escaped_text = en_text.replace("'", "''")
            values.append(f"('{source_text}', '{escaped_text}')")

    insert_values = ",".join(values)
    query = f"INSERT OR REPLACE INTO quests (ja, en) VALUES {insert_values};"

    db_query(query)

    # Story So Far worksheet
    values = []
    for i, _ in enumerate(ws_story_so_far, start=2):
        source_text = ws_story_so_far.cell(row=i, column=1).value
        deepl_text = ws_story_so_far.cell(row=i, column=2).value
        fixed_en_text = ws_story_so_far.cell(row=i, column=3).value

        if source_text and fixed_en_text:
            escaped_text = fixed_en_text.replace("'", "''")
            values.append(f"('{source_text}', '{escaped_text}')")
        elif source_text and deepl_text:
            escaped_text = deepl_text.replace("'", "''")
            values.append(f"('{source_text}', '{escaped_text}')")

    insert_values = ",".join(values)
    query = f"INSERT OR REPLACE INTO story_so_far_template (ja, en) VALUES {insert_values};"

    db_query(query)


def read_glossary_and_import(data: str) -> None:
    decoded_data = data.decode("utf-8")

    query_list = []
    glossary = [x for x in decoded_data.split("\n") if x]
    for record in glossary:
        ja, en = record.split(",", 1)

        escaped_value = en.replace("'", "''")
        query_value = f"('{ja}', '{escaped_value}')"

        query_list.append(query_value)

    drop_query = "DELETE FROM glossary;"
    db_query(drop_query)

    insert_values = ",".join(query_list)
    insert_query = f"INSERT OR REPLACE INTO glossary (ja, en) VALUES {insert_values};"
    db_query(insert_query)


def download_file(url: str) -> requests.models.Response:
    """Submits a GET request to a url with a retry count of 3.

    :param url: Request URL.
    :returns: A requests response object. On exception, will raise a
        SystemExit().
    """
    filename = url.split("/")[-1:][0]
    max_retries = 3
    retries = 0

    while retries < max_retries:
        try:
            # disabling TLS in general is an awful practice, but consumers of this application
            # are not technical and troubleshooting certificate issues with users is both
            # time consuming and exhausting. these SSL errors are primarily occurring specifically
            # with github.
            urllib3.disable_warnings()
            response = requests.get(url, verify=False)
            response.raise_for_status()
            break
        except Exception as e:
            retries += 1
            if retries < max_retries:
                log.warning(f"Error: {e}")
                log.warning(
                    f"(Retry: {retries}/{max_retries}) Error downloading file {filename}. ",
                    "Sleeping for 3 seconds and trying again...",
                )
                time.sleep(3)
                continue
            else:
                log.error(f"Error: {e}")
                log.error(f"(Retry: {retries}/{max_retries}) Max retries reached. Unable to download file {filename}.")
                log.info(
                    "\nIf connectivity issues persist, you can temporarily disable updates and try again later:\n"
                    '1) Check the "Disable Updates" checkbox in the launcher (or remove the -d flag)\n'
                    '2) Uncheck the "Update Game Files" checkbox in the launcher (or add the -u flag)\n\n'
                    "This is not a dqxclarity issue and should not be reported as such. You can check:\n"
                    "1) GitHub status page (https://www.githubstatus.com/)\n"
                    "2) Check your DNS settings (especially if the error says it failed to resolve)\n"
                    "3) Check that your VPN isn't blocking GitHub requests or is responding too slowly\n"
                    "4) Make sure you have a consistent internet connection. Bad WiFi?\n"
                    "5) If the error mentions SSL, try running Windows updates.\n"
                    "\nSCROLL UP AND READ!"
                )
                input("Press enter to exit.")
                sys.exit(1)

    return response


def import_name_overrides() -> None:
    """Adds the user's custom name overrides to the database if the file
    exists."""
    override_file = get_project_root("misc_files/name_overrides.json")

    if os.path.exists(override_file):
        with open(override_file, encoding="utf-8") as f:
            try:
                data = json.loads(f.read())
            except json.decoder.JSONDecodeError:
                log.error("name_overrides.json does not contain valid json; ignoring file.")
                return

            log.info("Importing player name overrides.")

            try:
                player_names = data["player_names"]
                mytown_names = data["mytown_names"]
            except KeyError as e:
                log.error(f"Missing key in name_overrides.json: {e}")
                log.warning("Ignoring overrides.")
                return

            if not player_names:
                log.info("No player names to override in name_overrides.json.")
            else:
                clean_custom_player_name_query = "DELETE FROM m00_strings WHERE file = 'local_player_names';"
                db_query(clean_custom_player_name_query)

                glossary_values = []
                m00_values = []
                for ja, en in player_names.items():
                    escaped_en = en.replace("'", "''")
                    glossary_values.append(f"('{ja}', '{escaped_en}')")
                    m00_values.append(f"('{ja}', '{escaped_en}', 'local_player_names')")

                # note: if a user removes override names from the file, they will continue to exist in the glossary
                # table until the user enables updates - which is what will clean up the glossary table.
                db_query(f"INSERT OR REPLACE INTO glossary (ja, en) VALUES {','.join(glossary_values)};")
                db_query(f"INSERT OR REPLACE INTO m00_strings (ja, en, file) VALUES {','.join(m00_values)};")

            if not mytown_names:
                log.info("No MyTown names to override in name_overrides.json.")
            else:
                clean_custom_mytown_name_query = "DELETE FROM m00_strings WHERE file = 'local_mytown_names';"
                db_query(clean_custom_mytown_name_query)

                mytown_values = []
                for ja, en in mytown_names.items():
                    escaped_en = en.replace("'", "''")
                    mytown_values.append(f"('{ja}', '{escaped_en}', 'local_mytown_names')")

                db_query(f"INSERT OR REPLACE INTO m00_strings (ja, en, file) VALUES {','.join(mytown_values)};")
    else:
        log.info("No name_overrides.json file found. Skipping import.")
        return
