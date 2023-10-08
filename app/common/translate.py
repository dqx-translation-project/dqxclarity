from common.db_ops import init_db
from common.errors import message_box
from common.lib import get_project_root, merge_jsons
from googleapiclient.discovery import build
from openpyxl import load_workbook

import configparser
import deepl
import json
import langdetect
import os
import pykakasi
import re
import shutil
import textwrap
import unicodedata


class Translate():
    service = None
    api_key = None
    region_code = None
    glossary = None

    def __init__(self):
        if Translate.service is None:
            self.user_settings = load_user_config()
            self.translation_settings = determine_translation_service()
            Translate.service = self.translation_settings["TranslateService"]
            Translate.api_key = self.translation_settings["TranslateKey"]
            Translate.region_code = self.translation_settings["RegionCode"]

        if Translate.glossary is None:
            with open(get_project_root("misc_files/glossary.csv"), encoding="utf-8") as f:
                strings = f.read()
                Translate.glossary = [ x for x in strings.split("\n") if x ]


    def deepl(self, text: list):
        region_code = Translate.region_code
        if region_code.lower() == "en":
            region_code = "en-us"
        translator = deepl.Translator(Translate.api_key)
        response = translator.translate_text(
            text=text,
            source_lang="ja",
            target_lang=region_code,
            preserve_formatting=True
        )
        text_results = []
        for result in response:
            text_results.append(result.text)
        return text_results


    def google(self, text: list):
        service = build("translate", "v2", developerKey=Translate.api_key)
        response = service.translations().list(source="ja", target="en", format="text", q=text).execute()
        text_results = []
        for result in response["translations"]:
            text_results.append(result["translatedText"])
        return text_results


    def __glossify(self, text):
        for record in Translate.glossary:
            k, v = record.split(",", 1)
            if v == "\"\"":  # check for glossary entries that have blank strings and re-assign
                v = ""
            text = text.replace(k, v)
        return text


    def __normalize_text(self, text: str) -> str:
        """"Normalize" text by only using latin alphabet.

        :param text: Text to normalize
        :returns: Normalized text.
        """
        return unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()


    def __swap_placeholder_tags(self, text: str, swap_back=False) -> str:
        if not swap_back:
            text = text.replace("<pc_hiryu>", "<&13_aaaaaaa>")
            text = text.replace("<cs_pchero_hiryu>", "<&13_aaaaaab>")
            text = text.replace("<cs_pchero_race>", "<&8_aaa>")
            text = text.replace("<cs_pchero>", "<13_aaaaaac>")
            text = text.replace("<kyodai_rel1>", "<&7_aa>")
            text = text.replace("<kyodai_rel2>", "<&7_ab>")
            text = text.replace("<kyodai_rel3>", "<&7_ac>")
            text = text.replace("<pc_hometown>", "<&8_aab>")
            text = text.replace("<pc_race>", "<&8_aac>")
            text = text.replace("<pc_rel1>", "<&7_ad>")
            text = text.replace("<pc_rel2>", "<&7_ae>")
            text = text.replace("<pc_rel3>", "<&7_af")
            text = text.replace("<kyodai>", "<&13_aaaaaac>")
            text = text.replace("<pc>", "<&13_aaaaaad>")
            text = text.replace("<client_pcname>", "<&13_aaaaaae>")
            text = text.replace("<heart>", "<&2a>")
            text = text.replace("<diamond>", "<&2b>")
            text = text.replace("<spade>", "<&2c>")
            text = text.replace("<clover>", "<&2d>")
            text = text.replace("<r_triangle>", "<&2e>")
            text = text.replace("<l_triangle>", "<&2f>")
            text = text.replace("<half_star>", "<&2g>")
            text = text.replace("<null_star>", "<&2h>")
            text = text.replace("<npc>", "<&13_aaaaaaf>")
            text = text.replace("<pc_syokugyo>", "<&13_aaaaaag>")
            text = text.replace("<pc_original>", "<&13_aaaaaah>")
            text = text.replace("<log_pc>", "<&13_aaaaaai>")
            text = text.replace("<1st_title>", "<&20_aaaaaaaaaaaaaa>")
            text = text.replace("<2nd_title>", "<&20_aaaaaaaaaaaaab>")
            text = text.replace("<3rd_title>", "<&20_aaaaaaaaaaaaac>")
            text = text.replace("<4th_title>", "<&20_aaaaaaaaaaaaad>")
            text = text.replace("<5th_title>", "<&20_aaaaaaaaaaaaae>")
            text = text.replace("<6th_title>", "<&20_aaaaaaaaaaaaaf>")
            text = text.replace("<7th_title>", "<&20_aaaaaaaaaaaaag>")
        else:
            text = text.replace("<&13_aaaaaaa>", "<pc_hiryu>")
            text = text.replace("<&13_aaaaaab>", "<cs_pchero_hiryu>")
            text = text.replace("<&8_aaa>", "<cs_pchero_race>")
            text = text.replace("<13_aaaaaac>", "<cs_pchero>")
            text = text.replace("<&7_aa>", "<kyodai_rel1>")
            text = text.replace("<&7_ab>", "<kyodai_rel2>")
            text = text.replace("<&7_ac>", "<kyodai_rel3>")
            text = text.replace("<&8_aab>", "<pc_hometown>")
            text = text.replace("<&8_aac>", "<pc_race>")
            text = text.replace("<&7_ad>", "<pc_rel1>")
            text = text.replace("<&7_ae>", "<pc_rel2>")
            text = text.replace("<&7_af", "<pc_rel3>")
            text = text.replace("<&13_aaaaaac>", "<kyodai>")
            text = text.replace("<&13_aaaaaad>", "<pc>")
            text = text.replace("<&13_aaaaaae>", "<client_pcname>")
            text = text.replace("<&2a>", "<heart>")
            text = text.replace("<&2b>", "<diamond>")
            text = text.replace("<&2c>", "<spade>")
            text = text.replace("<&2d>", "<clover>")
            text = text.replace("<&2e>", "<r_triangle>")
            text = text.replace("<&2f>", "<l_triangle>")
            text = text.replace("<&2g>", "<half_star>")
            text = text.replace("<&2h>", "<null_star>")
            text = text.replace("<&13_aaaaaaf>", "<npc>")
            text = text.replace("<&13_aaaaaag>", "<pc_syokugyo>")
            text = text.replace("<&13_aaaaaah>", "<pc_original>")
            text = text.replace("<&13_aaaaaai>", "<log_pc>")
            text = text.replace("<&20_aaaaaaaaaaaaaa>", "<1st_title>")
            text = text.replace("<&20_aaaaaaaaaaaaab>", "<2nd_title>")
            text = text.replace("<&20_aaaaaaaaaaaaac>", "<3rd_title>")
            text = text.replace("<&20_aaaaaaaaaaaaad>", "<4th_title>")
            text = text.replace("<&20_aaaaaaaaaaaaae>", "<5th_title>")
            text = text.replace("<&20_aaaaaaaaaaaaaf>", "<6th_title>")
            text = text.replace("<&20_aaaaaaaaaaaaag>", "<7th_title>")

        return text


    def __wrap_text(self, text: str, width: int, max_lines=None) -> str:
        """Wrap text to n characters per line."""
        return textwrap.fill(text, width=width, max_lines=max_lines, replace_whitespace=False)


    def __add_line_endings(self, text: str) -> str:
        """Adds <br> flags every 3 lines to a string. Used to break up the text
        in a dialog window.

        :param text: Text to add the <br> tags to.
        :returns: A new string with the text broken up by <br> tags.
        """
        count_list = [ i for i in range(3, 500, 4) ] # 500 is arbitrary, but we should never hit this.
        split_text = text.split("\n")
        try:
            for i in count_list:
                _ = split_text[i]
                split_text.insert(i, "<br>")
        except IndexError:
            split_text = [ x for x in split_text if x ]
            output = "\n".join(split_text)
            return output


    def translate(self, text: list):
        """Translates a list of strings, passing them through our glossary
        first.

        :param text: List of text strings to be translated.
        :returns: A translated list of strings in the same order they
            were given.
        """
        count = 0
        for i in text:
            text[count] = self.__glossify(i)
            count += 1
        if Translate.service == "deepl":
            return self.deepl(text)
        if Translate.service == "google":
            return self.google(text)
        return None


    def sanitize_and_translate(self, text: str, wrap_width: int, max_lines=None, add_brs=True):
        """Sanitizes different tags and symbols, then translates the string.

        :param text: String to be translated.
        :param wrap_width: How many characters the returning string
                should contain per line.
        :param max_lines: The maximum amount of lines to return. Extra
                lines are truncated with "..."
        :param add_brs: Whether to inject "<br>" every three lines to
                break up text. Used for dialog mainly.
        """
        if found := self.__search_excel_workbook(text):
            return found

        # manage our own line endings later
        output = text.replace("<br>", "　")

        # remove any tag alignments
        alignments = ["<center>", "<right>", "<left>"]
        for alignment in alignments:
            output = output.replace(alignment, "")

        # remove any other oddities that don't look great in english
        oddities = ["「", "…"]
        for oddity in oddities:
            output = output.replace(oddity, "")

        # remove the full width space that starts on a new line
        output = output.replace("\n　", "　")

        # replace any <color*> tags with & as they are part of the string
        output = output.replace("<color_", "<&color_")

        name_tags = ["<pc>", "<cs_pchero>", "<kyodai>"]

        # removes all of the honorifics added at the end of the tags
        honorifics = ["さま", "君", "どの", "ちゃん", "くん", "様", "さーん", "殿", "さん",]
        for tag in name_tags:
            for honorific in honorifics:
                output = output.replace(f"{tag}{honorific}", tag)

        # replace all variable name tags that expand to other text
        output = self.__swap_placeholder_tags(output)

        # pass string through our glossary to replace any common words
        output = self.__glossify(output)

        # re-assign this string. this is now our "pristine" string we'll be using later.
        pristine_str = output

        # get the text to translate, splitting on all tags that don't start with % or &
        tag_re = re.compile("(<[^%&]*?>)")
        select_re = re.compile(r"(<select.*>)")
        str_split = [ x for x in re.split(tag_re, output) if x ]

        count = 0
        str_attrs = {}

        # iterate over each string, handling based on condition
        for str in str_split:
            if not re.match(tag_re, str):

                # sole new lines need to stay where they are.
                if str == "\n":
                    continue

                # capture position of the string and replace with placeholder text
                pristine_str = pristine_str.replace(str, f"<replace_me_index_{count}>")

                # <select*> lists always start with their first entry being a newline.
                # if we see this, look back one index to see if we're inside a select tag.
                if str.startswith("\n"):
                    lookback = str_split.index(str) - 1
                    if re.match(select_re, str_split[lookback]):
                        str_attrs[count] = {
                            "text": str,
                            "is_list": True,
                            "prepend_newline": False,
                            "append_newline": False,
                        }
                        count += 1
                        continue

                # capture how the newline was originally placed
                append_newline = False
                if str.endswith("\n"):
                    append_newline = True

                prepend_newline = False
                if str.startswith("\n"):
                    prepend_newline = True

                str = str.replace("\n", "")

                str_attrs[count] = {
                    "text": str,
                    "is_list": False,
                    "prepend_newline": prepend_newline,
                    "append_newline": append_newline,
                }

                count += 1

        # translate our list of strings
        to_translate = []
        count = 0
        for str in str_attrs:
            to_translate.append(str_attrs[count]["text"])
            count += 1
        translated_list = self.translate(text=to_translate)

        # update our str_attrs dict with the new, translated string
        count = 0
        for str in translated_list:
            str_attrs[count]["text"] = str
            count += 1

        count = 0
        # search for any weird space usage and remove it.
        # this comes from deepl and are all scenarios that have been seen with
        # translations coming back from machine translation.
        for _ in str_attrs:
            str_text = str_attrs[count]["text"]
            str_text = str_text.replace("　 ", " ")
            str_text = str_text.replace("　", " ")
            str_text = str_text.replace("  ", "")

            updated_str = self.__normalize_text(str_text)
            updated_str = updated_str.replace("<&color_", "<color_")  # put our color tag back.

            if str_attrs[count]["is_list"]:
                # select lists will always have more than 1 entry..
                # leave selection lists alone. please don't fuck this up, deepl
                updated_str = self.__swap_placeholder_tags(updated_str, swap_back=True)

                # deepl occasionally indents our list lines.. even though they weren't originally indented
                updated_str = updated_str.replace("\n ", "\n")
                pristine_str = pristine_str.replace(f"<replace_me_index_{count}>", updated_str)

            else:
                # wrap the text and inject <br>'s to break the text up
                updated_str = self.__wrap_text(updated_str, width=wrap_width, max_lines=max_lines)
                updated_str = self.__swap_placeholder_tags(updated_str, swap_back=True)

                if add_brs:
                    updated_str = self.__add_line_endings(updated_str)
                if str_attrs[count]["prepend_newline"]:
                    updated_str = "\n" + updated_str
                if str_attrs[count]["append_newline"]:
                    updated_str += "\n"

                pristine_str = pristine_str.replace(f"<replace_me_index_{count}>", updated_str)

            count += 1

        return pristine_str


    def __search_excel_workbook(self, text: str):
        """Searches the merge.xlsx workbook for a string in the JP Text column.
        If there's a match and the string "BAD STRING" is found in the Notes
        column, this returns the contents in the "Fixed English Text" column.
        This fixes instances of text where machine translation completely
        screwed up the text and caused the game to have issues.

        :param text: String to search
        :returns: Returns either the English text or None if no match
            was found.
        """
        file = get_project_root("misc_files/merge.xlsx")

        if os.path.exists(file):
            wb = load_workbook(file)
            ws_dialogue = wb["Dialogue"]
            for rowNum in range(2, ws_dialogue.max_row + 1):
                jp_string = str(ws_dialogue.cell(row=rowNum, column=1).value)
                bad_string_col = str(ws_dialogue.cell(row=rowNum, column=4).value)
                if (jp_string in text) and ("BAD STRING" in bad_string_col):
                    return str(ws_dialogue.cell(row=rowNum, column=3).value)
        return None


def load_user_config():
    """Returns a user's config settings. If the config doesn't exist, a default
    config is generated. If the user's config is missing values, we back up the
    old config and generate a new default one for them.

    :returns: Dict of config.
    """
    filename = get_project_root("user_settings.ini")
    base_config = configparser.ConfigParser()
    base_config["translation"] = {
        "enabledeepltranslate": False,
        "deepltranslatekey": "",
        "enablegoogletranslate": False,
        "googletranslatekey": "",
        "regioncode": "en",
    }
    base_config["config"] = {"installdirectory": ""}

    def create_base_config():
        with open(filename, "w+") as configfile:
            base_config.write(configfile)

    # Create the config if it doesn't exist
    if not os.path.exists(filename):
        create_base_config()

    # Verify the integrity of the config. If a key is missing,
    # trigger user_config_state and create a new one, backing
    # up the old config.
    user_config = configparser.ConfigParser()
    user_config_state = 0
    user_config.read(filename)
    for section in base_config.sections():
        if section not in user_config.sections():
            user_config_state = 1
            break
        for key, val in base_config.items(section):
            if key not in user_config[section]:
                user_config_state = 1
                break

    # Notify user their config is busted
    if user_config_state == 1:
        shutil.copyfile(filename, f"{filename}.invalid")
        create_base_config()
        message_box(
            title="New config created",
            message=f"We found a missing config value in your {filename}.\n\nYour old config has been renamed to {filename}.invalid in case you need to reference it.\n\nPlease relaunch dqxclarity.",
            exit_prog=True,
        )

    config_dict = {}
    good_config = configparser.ConfigParser()
    good_config.read(filename)
    for section in good_config.sections():
        config_dict[section] = {}
        for key, val in good_config.items(section):
            config_dict[section][key] = val

    return config_dict


def update_user_config(section: str, key: str, value: str, filename="user_settings.ini"):
    """Updates an existing configuration option in a user's config.

    :param section: Section of the config
    :param key: Key in the section's config
    :param value: Value of the key
    :param filename: Filename of the user's config settings.
    """
    config = configparser.ConfigParser()
    config.read(filename)
    config.set(section, key, value)
    with open(filename, "w+") as configfile:
        config.write(configfile)


def determine_translation_service():
    """Parses the user_settings file to get information needed to make
    translation calls."""
    config = load_user_config()
    enabledeepltranslate = eval(config["translation"]["enabledeepltranslate"])
    deepltranslatekey = config["translation"]["deepltranslatekey"]
    enablegoogletranslate = eval(config["translation"]["enablegoogletranslate"])
    googletranslatekey = config["translation"]["googletranslatekey"]
    regioncode = config["translation"]["regioncode"]

    reiterate = "Either open the user_settings.ini file in Notepad or use the API settings button in the DQXClarity launcher to set it up."

    if enabledeepltranslate and enablegoogletranslate:
        message_box(
            title="Too many translation services enabled",
            message=f"Only enable one translation service. {reiterate}\n\nCurrent values:\n\nenabledeepltranslate: {enabledeepltranslate}\nenablegoogletranslate: {enablegoogletranslate}",
            exit_prog=True,
        )

    if enabledeepltranslate and deepltranslatekey == "":
        message_box(
            title="No DeepL key specified",
            message=f"DeepL is enabled, but no key was provided. {reiterate}",
            exit_prog=True,
        )

    if enablegoogletranslate and googletranslatekey == "":
        message_box(
            title="No Google API key specified",
            message=f"Google API is enabled, but no key was provided. {reiterate}",
            exit_prog=True,
        )

    dic = {}
    if enabledeepltranslate:
        dic["TranslateService"] = "deepl"
        dic["TranslateKey"] = deepltranslatekey
    elif enablegoogletranslate:
        dic["TranslateService"] = "google"
        dic["TranslateKey"] = googletranslatekey

    dic["RegionCode"] = regioncode

    return dic


def query_string_from_file(text: str, file: str) -> str:
    """Searches for a string from the specified json file and either returns
    the string or returns False if no match found.

    text: The text to search
    file: The name of the file (leave off the file extension)
    """
    misc_files = get_project_root("misc_files")
    data = read_json_file(misc_files + "/" + file + ".json")

    for item in data:
        key, value = list(data[item].items())[0]
        if re.search(f"^{text}+$", key):
            if value:
                return value


def clean_up_and_return_items(text: str) -> str:
    """Cleans up unnecessary text from item strings and searches for the name
    in items.json.

    Used specifically for the quest window.
    """
    misc_files = get_project_root("misc_files")
    quest_rewards = merge_jsons([
        f"{misc_files}/subPackage41Client.win32.json",
        f"{misc_files}/subPackage05Client.json",
        f"{misc_files}/custom_quest_rewards.json"
    ])
    line_count = text.count("\n")
    sanitized = re.sub("男は ", "", text)  # remove boy reference from start of string
    sanitized = re.sub("女は ", "", sanitized)  # remove girl reference from start of string
    sanitized = re.sub("男は　", "", sanitized)  # remove boy reference from start of string (fullwidth space)
    sanitized = re.sub("女は　", "", sanitized)  # remove girl reference from start of string (fullwidth space)
    sanitized = re.sub("！　と頼まれた。", "と頼まれた。", sanitized)  # remove annoying 'asked to' string
    final_string = ""
    for item in sanitized.split("\n"):
        quantity = ""
        no_bullet = re.sub(r"(^\・)", "", item)
        points = no_bullet[6:18]
        if no_bullet.endswith("こ"):
            quantity = "(" + unicodedata.normalize("NFKC", no_bullet[-3:-1]) + ")"
            quantity = re.sub(" ", "", quantity)
        if no_bullet.endswith("他"):
            bad_strings = ["必殺技を覚える", "入れられるよう"]
            if any(string in no_bullet for string in bad_strings):
                quantity = ""
            else:
                quantity = "(1)"
        no_bullet = re.sub("(　　.*)", "", no_bullet)
        if no_bullet in quest_rewards:
            value = quest_rewards.get(no_bullet)
            if value:
                value_length = len(value)
                quant_length = len(quantity)
                byte_count = len(value.encode('utf-8'))
                num_spaces = 31 - value_length - quant_length - ((byte_count - value_length)//2)
                if "・" in item:
                    if line_count == 0:
                        return "・" + value + (" " * num_spaces) + quantity
                    else:
                        final_string += "・" + value + (" " * num_spaces) + quantity + "\n"
                else:
                    if line_count == 0:
                        return value + (" " * num_spaces) + quantity
                    else:
                        final_string += value + (" " * num_spaces) + quantity + "\n"
        else:
            if line_count == 0:
                if "討伐ポイント" in item:
                    return "・" + "Experience Points" + points
                else:
                    return text
            else:
                final_string += item + "\n"
    return final_string.rstrip()


def detect_lang(text: str) -> bool:
    """Detects if the language is Japanese or not.

    Returns bool.
    """
    sanitized = re.sub("<.+?>", "", text)
    sanitized = re.sub("\n", "", sanitized)

    try:
        if langdetect.detect(sanitized) == "ja":
            return True
    except langdetect.lang_detect_exception.LangDetectException:  # Could not detect language
        return False


def read_json_file(file):
    with open(file, encoding="utf-8") as json_data:
        return json.loads(json_data.read())


def convert_into_eng(word: str) -> str:
    """Uses the pykakasi library to phonetically convert a Japanese word
    (usually a name) into English.

    :param word: Word to convert.
    :returns: Returns up to a 10 character name in English.
    """
    kks = pykakasi.kakasi()
    invalid_chars = ["[", "]", "[", "(", ")", "\\", "/", "*", "_", "+", "?", "$", "^", '"']
    misc_files = get_project_root("misc_files")
    player_names = merge_jsons([f"{misc_files}/custom_player_names.json", f"{misc_files}/custom_npc_names.json"])
    interpunct_count = word.count("・")
    word_len = len(word)
    bad_word = False

    if any(char in word for char in invalid_chars):
        return word
    else:
        romaji_name = ""
        if word in player_names:
            value = player_names.get(word)
            if value:
                romaji_name = value
            return romaji_name[0:10]
        else:
            if word_len < 7:
                for char in word:
                    num = ord(char)
                    if num not in (list(range(12353, 12430)) + [12431] + list(range(12434,12436)) + list(range(12449,12526)) + [12527] + list(range(12530,12533)) + list(range(12539,12541)) + [65374]):
                        bad_word = True
                        return word
                if bad_word != True:
                    result = kks.convert(word)
                    for word in result:
                        romaji_name = romaji_name + word["hepburn"]
                    romaji_name = romaji_name.title()
                    romaji_name = romaji_name.replace("・", "")
                    if romaji_name == "":
                        romaji_name = "." * interpunct_count
                    return romaji_name[0:10]
            else:
                return word


def get_player_name() -> tuple:
    """Queries the player and sibling name from the database.

    Returns a tuple of (player_name, sibling_name).
    """
    conn, cursor = init_db()

    player_query = "SELECT name FROM player WHERE type = 'player'"
    sibling_query = "SELECT name FROM player WHERE type = 'sibling'"

    results = cursor.execute(player_query)
    player = results.fetchone()[0]

    results = cursor.execute(sibling_query)
    sibling = results.fetchone()[0]

    conn.close()

    return (player, sibling)
